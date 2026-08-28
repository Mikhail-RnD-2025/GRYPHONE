# -*- coding: utf-8 -*-
"""
import_from_excel.py — импорт камер из Excel файла.

ПОЛНАЯ ПОДДЕРЖКА ВСЕХ ПОЛЕЙ КАМЕРЫ (как в cameras.json):
  ┌────────────┬─────────┬──────────────┬──────────────────────┐
  │ Поле       │ Тип     │ Обязательно  │ По умолчанию         │
  ├────────────┼─────────┼──────────────┼──────────────────────┤
  │ id         │ str     │ ДА           │ —                    │
  │ name       │ str     │ нет          │ = id                 │
  │ main_url   │ str     │ ДА           │ —                    │
  │ sub_url    │ str     │ нет          │ ""                   │
  │ enabled    │ bool    │ нет          │ true                 │
  │ comment    │ str     │ нет          │ ""                   │
  │ audio      │ bool    │ нет          │ true                 │
  │ location   │ str     │ нет          │ ""                   │
  └────────────┴─────────┴──────────────┴──────────────────────┘

Автоматически определяет колонки по заголовкам (регистронезависимо).
Поддерживает русский и английский названия колонок.

Запуск через API: см. endpoint /api/cameras/import-excel
Запуск через CLI: python import_from_excel.py cameras.xlsx
"""
import sys
import json
import sqlite3
import re
import logging
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ Библиотека openpyxl не установлена!")
    print("   Установите: pip install openpyxl")
    sys.exit(1)

logger = logging.getLogger(__name__)

# ============================================================
# ПОЛНАЯ СТРУКТУРА ПОЛЕЙ КАМЕРЫ
# ============================================================
# Каждое поле описано явно: тип, обязательность, значение по умолчанию.
# Это гарантирует, что все поля из cameras.json обрабатываются корректно.
CAMERA_FIELDS = {
    'id':       {'type': 'str',  'required': True,  'default': None},
    'name':     {'type': 'str',  'required': False, 'default': ''},
    'main_url': {'type': 'str',  'required': True,  'default': None},
    'sub_url':  {'type': 'str',  'required': False, 'default': ''},
    'enabled':  {'type': 'bool', 'required': False, 'default': True},
    'comment':  {'type': 'str',  'required': False, 'default': ''},
    'audio':    {'type': 'bool', 'required': False, 'default': True},
    'location': {'type': 'str',  'required': False, 'default': ''},
}

# Маппинг возможных названий колонок (нижний регистр).
# Поддерживает русский и английский варианты.
COLUMN_MAPPING = {
    'id': [
        'id', 'ID', 'идентификатор', 'id камеры', 'camera id',
        'номер', 'код', 'code',
    ],
    'name': [
        'name', 'имя', 'название', 'камера', 'camera name',
        'название камеры', 'имя камеры',
    ],
    'main_url': [
        'main_url', 'main url', 'основной_url', 'основной url',
        'url', 'ссылка', 'основная ссылка', 'поток', 'stream',
        'rtsp', 'main', 'поток основной', 'основной поток',
    ],
    'sub_url': [
        'sub_url', 'sub url', 'дополнительный_url', 'дополнительный url',
        'субпоток', 'доп ссылка', 'дополнительная ссылка', 'доп поток',
        'дополнительный поток', 'второй поток', 'поток 2',
    ],
    'enabled': [
        'enabled', 'включена', 'статус', 'активна', 'активен',
        'вкл', 'выкл', 'активная', 'включить',
    ],
    'comment': [
        'comment', 'комментарий', 'описание', 'примечание',
        'заметка', 'доп инфо', 'доп. информация',
    ],
    'audio': [
        'audio', 'звук', 'аудио', 'со звуком', 'звук вкл',
        'аудио вкл', 'звук?', 'аудио?',
    ],
    'location': [
        'location', 'расположение', 'местоположение', 'адрес',
        'место', 'где установлена', 'помещение',
    ],
}


# ============================================================
# ФУНКЦИИ ОЧИСТКИ ДАННЫХ
# ============================================================
def clean_str(value: Any) -> str:
    """
    Очищает строковое значение от лишних пробелов.

    Обрабатывает особенность исходных данных:
      " 210-P-GAVw-001 " → "210-P-GAVw-001"
    """
    if value is None:
        return ''
    return str(value).strip()


def clean_name(value: Any) -> str:
    """
    Очищает имя камеры от лишних пробелов и дефиса в начале.

    Обрабатывает особенность исходных данных:
      " -210-P-GAVw-001 " → "210-P-GAVw-001"
    """
    cleaned = clean_str(value)
    # Убираем дефис в начале, если он есть (артефакт экспорта)
    if cleaned.startswith('-'):
        cleaned = cleaned[1:].strip()
    return cleaned


def clean_url(value: Any) -> str:
    """
    Очищает RTSP-ссылку от лишних пробелов (включая внутри).

    Обрабатывает особенность исходных данных:
      "rtsp://...?channel=1 &subtype=0 " → "rtsp://...?channel=1&subtype=0"
    """
    if value is None:
        return ''
    # Убираем пробелы в начале/конце и все внутренние пробелы
    return str(value).strip().replace(' ', '')


def parse_bool(value: Any, default: bool = True) -> bool:
    """
    Парсит булево значение из разных форматов.

    Поддерживаемые форматы:
      - bool: True/False
      - int/float: 1/0, 1.0/0.0
      - str: "true"/"false", "да"/"нет", "1"/"0", "вкл"/"выкл"

    Если значение пустое или не распознано — возвращает `default`.
    """
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)

    value_str = clean_str(value).lower()
    if not value_str:
        return default

    true_values = ['true', '1', 'да', 'yes', 'вкл', 'включена', 'активна', 'включено']
    false_values = ['false', '0', 'нет', 'no', 'выкл', 'выключена', 'неактивна', 'выключено']

    if value_str in true_values:
        return True
    if value_str in false_values:
        return False
    return default


# ============================================================
# ОПРЕДЕЛЕНИЕ КОЛОНОК
# ============================================================
def detect_columns(headers: List[str]) -> Tuple[Dict[str, int], List[str]]:
    """
    Определяет индексы колонок по заголовкам.

    Проходит по всем полям из CAMERA_FIELDS и ищет соответствующие
    колонки в заголовках. Регистр не учитывается.

    Returns:
        Tuple[Dict[str, int], List[str]]:
          - маппинг поле → индекс колонки
          - список найденных полей для логирования
    """
    mapping = {}
    found_fields = []
    headers_lower = [h.strip().lower() if h else '' for h in headers]

    for field in CAMERA_FIELDS.keys():
        possible_names = COLUMN_MAPPING.get(field, [field])
        for idx, header in enumerate(headers_lower):
            if header in [name.lower() for name in possible_names]:
                mapping[field] = idx
                found_fields.append(f"{field} (колонка {idx+1}: '{headers[idx]}')")
                break

    return mapping, found_fields


# ============================================================
# ЧТЕНИЕ EXCEL ФАЙЛА
# ============================================================
def read_excel_file(filepath: Path) -> List[Dict[str, Any]]:
    """
    Читает Excel файл и возвращает список камер со всеми 8 полями.

    Каждый элемент списка содержит все поля из CAMERA_FIELDS:
      id, name, main_url, sub_url, enabled, comment, audio, location
    """
    if not filepath.exists():
        raise ValueError(f"Файл не найден: {filepath}")

    print(f"📂 Чтение файла: {filepath}")

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    if ws is None:
        raise ValueError("Не удалось прочитать активный лист")

    # Читаем заголовки (первая строка)
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(cell.value if cell.value is not None else '')

    print(f"📋 Заголовки колонок: {headers}")

    # Определяем маппинг колонок
    col_map, found_fields = detect_columns(headers)
    print(f"🔍 Определены колонки:")
    for f in found_fields:
        print(f"   • {f}")

    # Проверяем обязательные поля
    required = [f for f, spec in CAMERA_FIELDS.items() if spec['required']]
    missing = [f for f in required if f not in col_map]
    if missing:
        raise ValueError(
            f"Не найдены обязательные колонки: {', '.join(missing)}. "
            f"Найдены: {', '.join(found_fields)}"
        )

    # Читаем данные
    cameras = []
    row_num = 1
    empty_rows = 0
    skipped_rows = 0

    for row in ws.iter_rows(min_row=2):
        row_num += 1
        values = [cell.value for cell in row]

        # Пропускаем полностью пустые строки
        if not any(v is not None for v in values):
            empty_rows += 1
            continue

        try:
            # Извлекаем каждое поле по маппингу.
            # Если колонка не найдена — используем значение по умолчанию.
            camera = {}
            for field, spec in CAMERA_FIELDS.items():
                if field in col_map:
                    raw_value = values[col_map[field]]
                    if spec['type'] == 'str':
                        if field == 'name':
                            camera[field] = clean_name(raw_value)
                        elif field in ('main_url', 'sub_url'):
                            camera[field] = clean_url(raw_value)
                        else:
                            camera[field] = clean_str(raw_value)
                    elif spec['type'] == 'bool':
                        camera[field] = parse_bool(raw_value, spec['default'])
                else:
                    # Колонка не найдена — используем значение по умолчанию
                    camera[field] = spec['default']

            # Валидация обязательных полей
            if not camera.get('id') or not camera.get('main_url'):
                print(f"⚠️  Строка {row_num}: пропущена (нет ID или URL)")
                skipped_rows += 1
                continue

            # Если имя пустое — используем ID
            if not camera['name']:
                camera['name'] = camera['id']

            # Если sub_url пустой или совпадает с main_url — оставляем пустым
            if camera['sub_url'] == camera['main_url']:
                camera['sub_url'] = ''

            cameras.append(camera)

        except Exception as e:
            print(f"⚠️  Строка {row_num}: ошибка парсинга - {e}")
            skipped_rows += 1
            continue

    wb.close()

    if not cameras:
        raise ValueError("Не найдено ни одной валидной камеры в файле")

    print(f"\n✅ Прочитано камер: {len(cameras)}")
    print(f"   Пропущено пустых строк: {empty_rows}")
    print(f"   Пропущено невалидных строк: {skipped_rows}")

    return cameras


# ============================================================
# СОЗДАНИЕ НАБОРОВ
# ============================================================
def create_sets_from_cameras(cameras: List[Dict]) -> Dict:
    """
    Автоматически создаёт наборы на основе префиксов ID камер.

    Логика группировки:
      - "210-P-GAVw-001" → группа "210"
      - "301A-P-GAVw-001" → группа "301A"
      - "phone", "test" → группа "other"
    """
    groups: Dict[str, List[str]] = {}

    for cam in cameras:
        cam_id = cam['id']
        parts = cam_id.split('-')
        group = parts[0] if len(parts) > 1 else 'other'
        if group not in groups:
            groups[group] = []
        groups[group].append(cam_id)

    # Функция расчёта размеров сетки
    def calc_grid(count: int) -> Tuple[int, int]:
        if count <= 4: return 2, 2
        elif count <= 9: return 3, 3
        elif count <= 16: return 4, 4
        elif count <= 25: return 5, 5
        elif count <= 36: return 6, 6
        elif count <= 49: return 7, 7
        else: return 8, 8

    sets_data = {'default_set': '', 'sets': {}}

    for group_id, cam_ids in groups.items():
        cols, rows = calc_grid(len(cam_ids))
        sets_data['sets'][group_id] = {
            'name': f'📹 {group_id}',
            'camera_ids': cam_ids,
            'max_columns': cols,
            'max_rows': rows,
            'aspect_ratio': '16:9',
        }

    # Общий набор "Все камеры"
    sets_data['sets']['all'] = {
        'name': '📦 Все камеры',
        'camera_ids': [c['id'] for c in cameras],
        'max_columns': 8,
        'max_rows': 8,
        'aspect_ratio': '16:9',
    }
    sets_data['default_set'] = 'all'

    return sets_data


# ============================================================
# СОХРАНЕНИЕ В БД
# ============================================================
def save_to_database(cameras: List[Dict], sets_data: Dict,
                     db_path: Path = Path(str(DATABASE_PATH))) -> None:
    """Сохраняет камеры и наборы в базу данных."""
    print(f"\n💾 Сохранение в БД: {db_path}")

    conn = sqlite3.connect(str(db_path))
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
    ''')

    cursor.execute(
        'INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)',
        ('cameras', json.dumps(cameras, ensure_ascii=False, indent=2))
    )
    print(f"✅ Сохранено камер: {len(cameras)}")

    cursor.execute(
        'INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)',
        ('sets', json.dumps(sets_data, ensure_ascii=False, indent=2))
    )
    print(f"✅ Сохранено наборов: {len(sets_data['sets'])}")

    conn.commit()
    conn.close()


# ============================================================
# СТАТИСТИКА
# ============================================================
def extract_ip(url: str) -> Optional[str]:
    """Извлекает IP из RTSP-ссылки."""
    match = re.search(r'//(?:[^@]+@)?([^:/]+)', url or '')
    return match.group(1) if match else None


def print_statistics(cameras: List[Dict], sets_data: Dict) -> None:
    """Выводит подробную статистику импорта."""
    print("\n" + "="*70)
    print("📊 СТАТИСТИКА ИМПОРТА")
    print("="*70)

    total = len(cameras)
    enabled = sum(1 for c in cameras if c['enabled'])
    with_audio = sum(1 for c in cameras if c['audio'])
    with_sub = sum(1 for c in cameras if c['sub_url'])
    with_location = sum(1 for c in cameras if c['location'])
    with_comment = sum(1 for c in cameras if c['comment'])

    print(f"📹 Всего камер: {total}")
    print(f"   ✅ Включено: {enabled}")
    print(f"   ❌ Выключено: {total - enabled}")
    print(f"   🔊 С аудио: {with_audio}")
    print(f"   🎥 С субпотоком: {with_sub}")
    print(f"   📍 С местоположением: {with_location}")
    print(f"   💬 С комментарием: {with_comment}")

    # Статистика по IP
    ips: Dict[str, int] = {}
    for cam in cameras:
        ip = extract_ip(cam['main_url'])
        if ip:
            ips[ip] = ips.get(ip, 0) + 1

    print(f"\n🌐 Уникальных IP: {len(ips)}")
    if len(ips) <= 20:
        for ip, count in sorted(ips.items(), key=lambda x: -x[1])[:10]:
            print(f"   {ip}: {count} камер")

    # Статистика по наборам
    print(f"\n📦 Наборы:")
    for set_id, s in sets_data['sets'].items():
        grid = f"{s['max_columns']}×{s['max_rows']}"
        print(f"   {s['name']}: {len(s['camera_ids'])} камер (сетка {grid})")

    print("="*70)


# ============================================================
# ГЛАВНАЯ ФУНКЦИЯ ИМПОРТА
# ============================================================
def import_excel(filepath: str, db_path: str = str(DATABASE_PATH)) -> Dict:
    """
    Главная функция импорта.

    Читает Excel, создаёт наборы, сохраняет в БД.
    Все 8 полей камеры обрабатываются корректно.
    """
    try:
        cameras = read_excel_file(Path(filepath))
        sets_data = create_sets_from_cameras(cameras)
        print_statistics(cameras, sets_data)
        save_to_database(cameras, sets_data, Path(db_path))

        print("\n✅ Импорт завершён успешно!")

        return {
            'success': True,
            'cameras_count': len(cameras),
            'sets_count': len(sets_data['sets']),
            'message': f"Импортировано {len(cameras)} камер в {len(sets_data['sets'])} наборов",
        }

    except Exception as e:
        error_msg = f"❌ Ошибка импорта: {e}"
        print(error_msg)
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'error': str(e),
            'message': error_msg,
        }


# ============================================================
# CLI
# ============================================================
if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Использование: python import_from_excel.py <файл.xlsx> [db_path]")
        print("\nПоддерживаемые поля в Excel:")
        print("  Обязательные: ID, main_url")
        print("  Опциональные: name, sub_url, enabled, comment, audio, location")
        sys.exit(1)

    filepath = sys.argv[1]
    db_path = sys.argv[2] if len(sys.argv) > 2 else str(DATABASE_PATH)

    result = import_excel(filepath, db_path)
    if not result['success']:
        sys.exit(1)
